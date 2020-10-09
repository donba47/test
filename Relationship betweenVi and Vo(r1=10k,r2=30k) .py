# -*- coding: utf-8 -*-
"""
Created on Sat Apr 18 11:25:36 2020

@author: petertw89
"""
import numpy as np
import matplotlib.pyplot as plt
from scipy import optimize
import scipy.stats as stats
import openpyxl as op

def f_1(x, A, B):


    return A*x + B  
plt.figure(figsize=(8,6),dpi=80)   
x=[]
y=[]
y3=[-8.51,-8.51,7.85,7.85]
x3=[-2.5,-2.25,2.25,2.5]
workbook=op.load_workbook('同向放大器數據.xlsx')
sheet=workbook.worksheets[0]
for i in range(4,12+1):
    x.append(sheet.cell(row=5,column=i).value)
    y.append(sheet.cell(row=6,column=i).value)

print(x,y)
plt.style.use("ggplot")     
plt.xlabel("Vi   (Volt)", fontweight = "bold",fontsize=14)
plt.ylabel("Vo   (Volt)", fontweight = "bold",fontsize=14)  

plt.title("Relationship between Vi and Vo(R1=10Kohm&R2=30Kohm)",
           fontweight = "bold",fontsize=16)        

plt.scatter(x,y,c = "m",s = 50)       
plt.axhline(y=0,c='black')
plt.axvline(x=0,c='black')                   
A2,B2=optimize.curve_fit(f_1, x, y)[0]
x2 = np.arange(-2,2,0.0001)
y2 = A2*x2 + B2
r=stats.pearsonr(x,y)
#plt.text(0.014,1.66, r'Correlation coefficient=%s'%(round(r[0],6)))
plt.scatter(x3,y3,c = "m",s = 50)  
print("Av=%s"%A2)
Avdb=20*np.log(np.abs(A2))
print("Av(dB)=%s"%Avdb)
plt.plot(x2, y2,"--",label="y=%s*x+%s"%(round(A2,9),round(B2,9))) 
plt.text(1,-5, r'Av=%s'%(round(A2,4)),fontsize=15)
plt.text(1,-6, r'Av|db=%s'%(round(Avdb,2)),fontsize=15)    
plt.legend(loc="best")
plt.savefig("Relationship between Vi and Vo(R1=10Kohm&R2=30Kohm)") 
plt.show()
plt.close() 

