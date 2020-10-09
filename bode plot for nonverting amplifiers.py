# -*- coding: utf-8 -*-
"""
Created on Mon Sep 28 16:28:57 2020

@author: petertw89
"""
from scipy import signal
import numpy as np
import matplotlib.pyplot as plt
from scipy import optimize
import scipy.stats as stats
import openpyxl as op
plt.figure(figsize=(10,10),dpi=80)   
y1=[]
x1=[10,20,30,50,100,200,300,500,1000,2000,3000,5000,10000,20000,30000,50000,100000,200000,300000,500000,1e6,2e6]
x2=[10,20,30,50,100,200,300,500,1000,2000,3000,5000,10000,20000,30000,50000,100000,200000,300000,500000,1e6,2e6]
y2=[]
f= []
plt.style.use("ggplot")     
workbook=op.load_workbook('同向放大器數據.xlsx')
sheet=workbook.worksheets[0]
for i in range(2,12+1):
    y1.append(sheet.cell(row=19,column=i).value)
    y2.append(sheet.cell(row=20,column=i).value)
for i in range(2,12+1):
    y1.append(sheet.cell(row=26,column=i).value)
    y2.append(sheet.cell(row=27,column=i).value)
print(y2,y1)
for i in y2:
    f.append(i*np.pi*2)


   

   
#plt.text(0.01,1.90, r'Correlation coefficient=%s'%(round(r[0],6)))
plt.subplot(211)
plt.subplots_adjust(wspace=0,hspace =0.35)
plt.subplots_adjust(left=None, bottom=None, right=None, top=None, wspace=None, hspace=None)
plt.semilogx(x1, y1)
plt.scatter(x1,y1)
plt.ylabel("Magnitude(dB)", fontweight = "bold",fontsize=14)
plt.xlabel("Frequency(Hz)", fontweight = "bold",fontsize=14)  
plt.title("Bode Diagram for inverting amplifiers",
           fontweight = "bold",fontsize=16)     
plt.legend(loc="best")

plt.subplot(212)
plt.semilogx(x2,y2)
plt.scatter(x2,y2)

plt.xlabel("Frequency(Hz)", fontweight = "bold",fontsize=14)
plt.ylabel("Phase(deg)", fontweight = "bold",fontsize=14)
plt.ylim(0,)  
plt.legend(loc="best")
plt.savefig("bode plot for inverting amplifiers") 
plt.show()
plt.close() 
